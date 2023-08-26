

import os
import os.path
import datetime

import customtkinter as ctk
import openpyxl
from openpyxl import Workbook




class App(ctk.CTk):
    def __init__(nmod):
        super().__init__()

        nmod.title("Poject Ideas")
        nmod.geometry("600x400")
        nmod.resizable(0, 0)
        nmod.grid_rowconfigure(0, weight=2)
        # nmod.grid_columnconfigure(1, weight=1)

        _datetime = datetime.datetime.now()
        _date = _datetime.strftime("%d/%m/%Y")
        # _time = ("%H:%M:%S %p")

        dat1=str("data/poject_idea.xlsx")
        check_dat1=os.path.isfile(dat1)

        idea_value = ctk.StringVar()
        date_value = ctk.StringVar()

        date_value.set(_date)

        if check_dat1==False:
            new_dat1=Workbook()
            new_dat1.save(dat1)

            dat1_file=openpyxl.load_workbook(dat1)
            dat1_sheet=dat1_file.active
            max_column_dat1=dat1_sheet.max_column
            max_row_dat1=dat1_sheet.max_row

            dat1_sheet["A1"].value="DATE"
            dat1_sheet["B1"].value="IDEA"
            dat1_sheet.column_dimensions["A"].width=15
            dat1_sheet.column_dimensions["B"].width=100
            dat1_file.save(dat1)



        def value_update():
            try:
                dat1_file=openpyxl.load_workbook(dat1)
                dat1_sheet=dat1_file.active
                max_column_dat1=dat1_sheet.max_column
                max_row_dat1=dat1_sheet.max_row

                column_name=dat1_sheet.cell(row=max_row_dat1+1, column=1)
                column_name.value=date_value.get()

                column_class=dat1_sheet.cell(row=max_row_dat1+1, column=2)
                column_class.value=idea_value.get()

                # print(date_value.get(), idea_value.get())
                nmod.log_textbox.configure(state="normal")
                nmod.log_textbox.insert("0.0", f"Sucessfuly Saved!\n{date_value.get()}   -   {idea_value.get()}\n",)
                nmod.log_textbox.tag_add("start", "1.0", "1.17")
                nmod.log_textbox.tag_config("start", foreground="#02fa44")
                nmod.log_textbox.configure(state="disabled")
                dat1_file.save(dat1)
            except:
                nmod.log_textbox.configure(state="normal")
                nmod.log_textbox.delete("1.0", "end")  # clears all logs from textbox
                nmod.log_textbox.insert("0.0", "Data File Error!\n")
                nmod.log_textbox.tag_add("start", "1.0", "1.16")
                nmod.log_textbox.tag_config("start", foreground="red")
                nmod.log_textbox.configure(state="disabled")


        def get_data():
            try:
                dat1_file=openpyxl.load_workbook(dat1)
                dat1_sheet=dat1_file.active
                max_column_dat1=dat1_sheet.max_column
                max_row_dat1=dat1_sheet.max_row
                nmod.log_textbox.configure(state="normal")
                nmod.log_textbox.delete("1.0", "end")  # clears all logs from textbox

                for row in range(2, max_row_dat1+1):
                    column1=dat1_sheet.cell(row=row, column=1)
                    column2=dat1_sheet.cell(row=row, column=2)
                    nmod.log_textbox.configure(state="normal")
                    log_data=(f"{column1.value}   -   {column2.value}\n")
                    nmod.log_textbox.insert("1.0", log_data)
                    nmod.log_textbox.configure(state="disabled")
            except:
                nmod.log_textbox.configure(state="normal")
                nmod.log_textbox.delete("1.0", "end")  # clears all logs from textbox
                nmod.log_textbox.insert("0.0", "Data File Error!\n")
                nmod.log_textbox .tag_add("start", "1.0", "1.16")
                nmod.log_textbox.tag_config("start", foreground="red")
                nmod.log_textbox.configure(state="disabled")


        dat1_file=openpyxl.load_workbook(dat1)
        dat1_sheet=dat1_file.active
        max_column_dat1=dat1_sheet.max_column 
        max_row_dat1=dat1_sheet.max_row


            

        nmod.time_label = ctk.CTkLabel(nmod, text="Date").place(x=50, y=22)

        nmod.time_entry = ctk.CTkEntry(nmod, textvariable=date_value, height=25, width=100)
        nmod.time_entry.place(x=50, y=50)

        nmod.idea_label = ctk.CTkLabel(nmod, text="Input your idea :D").place(x=200, y=22)

        nmod.idea_entry = ctk.CTkEntry(nmod, textvariable=idea_value, height=25, width=200, fg_color="#000000", text_color="#ffffff")
        nmod.idea_entry.place(x=200, y=50)

        nmod.save_button = ctk.CTkButton(nmod, text="Save", border_width=None, height=25, width=100, text_color=("gray10", "#DCE4EE"), command=value_update)
        nmod.save_button.place(x=450, y=50)

        nmod.log_textbox = ctk.CTkTextbox(nmod, wrap="none", height=225, width=500, fg_color=None, corner_radius=5, border_spacing=5)
        nmod.log_textbox.pack(pady=(130,0))
        nmod.log_textbox.configure(state="disabled")

        nmod.idea_button = ctk.CTkButton(nmod, text="View Ideas", border_width=None, height=25, width=100, text_color=("gray10", "#DCE4EE"), command=get_data)
        nmod.idea_button.place(x=250, y=100)


        # nmod.log_textbox.insert((0:0), "")


        




if __name__=="__main__":
    app=App()
    app.mainloop()